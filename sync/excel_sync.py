"""
Reads an Excel file and syncs all sheets into the PostgreSQL database.
Each sheet is treated as an independent table. Headers come from row 1.

Usage:
    python excel_sync.py path/to/file.xlsx
    python excel_sync.py          # uses EXCEL_PATH from .env or ../fake_data.xlsx
"""
import openpyxl
import psycopg2
import psycopg2.extras
import json
import logging
import os
import sys
import time
from datetime import datetime
from logging.handlers import RotatingFileHandler
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), '../backend/.env'))

DATABASE_URL = os.getenv("DATABASE_URL")

_LOG_PATH = os.getenv(
    "SYNC_LOG_PATH",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'sync.log')
)

logging.basicConfig(
    level=logging.INFO,
    format='%(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        RotatingFileHandler(_LOG_PATH, maxBytes=1_000_000, backupCount=3, encoding='utf-8'),
    ]
)
logger = logging.getLogger(__name__)


def serialize_value(val):
    """Convert Excel cell values to JSON-serializable types."""
    if val is None:
        return None
    if hasattr(val, 'isoformat'):  # datetime / date
        return val.isoformat()
    if isinstance(val, float) and (val != val):  # NaN check
        return None
    return val


def sync_excel(file_path: str):
    timestamp = datetime.now().strftime('%H:%M:%S')
    logger.info(f"[{timestamp}] Syncing: {file_path}")

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except FileNotFoundError:
        logger.info(f"  ERROR: File not found: {file_path}")
        return False
    except Exception as e:
        logger.info(f"  ERROR: Could not open file: {e}")
        return False

    try:
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
    except Exception as e:
        logger.info(f"  ERROR: Could not connect to database: {e}")
        return False

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))

            if not all_rows or not any(all_rows[0]):
                logger.info(f"  Skipping '{sheet_name}': empty or no headers")
                continue

            headers = [str(h).strip() for h in all_rows[0] if h is not None]
            data_rows = [
                r for r in all_rows[1:]
                if any(v is not None for v in r)
            ]

            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO sheets (name, display_name, columns, last_synced_at)
                    VALUES (%s, %s, %s::jsonb, NOW())
                    ON CONFLICT (name) DO UPDATE SET
                        columns = EXCLUDED.columns,
                        last_synced_at = EXCLUDED.last_synced_at
                    RETURNING id
                """, (sheet_name, sheet_name, json.dumps(headers)))
                sheet_id = cur.fetchone()["id"]

                cur.execute("DELETE FROM sheet_data WHERE sheet_id = %s", (sheet_id,))

                for row in data_rows:
                    row_dict = {
                        headers[i]: serialize_value(row[i] if i < len(row) else None)
                        for i in range(len(headers))
                    }
                    cur.execute(
                        "INSERT INTO sheet_data (sheet_id, row_data) VALUES (%s, %s::jsonb)",
                        (sheet_id, json.dumps(row_dict))
                    )

            conn.commit()
            logger.info(f"  '{sheet_name}': {len(data_rows)} rows, {len(headers)} columns")

        logger.info(f"  Sync complete.\n")
        return True

    except Exception as e:
        conn.rollback()
        logger.info(f"  ERROR during sync: {e}")
        return False
    finally:
        conn.close()


def run_once(file_path: str):
    sync_excel(file_path)


def run_loop(file_path: str, interval_seconds: int = 5):
    """Run sync continuously, every interval_seconds. Resilient to failures."""
    logger.info(f"Starting sync loop — interval: {interval_seconds}s — file: {file_path}\n")
    while True:
        sync_excel(file_path)
        time.sleep(interval_seconds)


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else os.getenv(
        "EXCEL_PATH",
        os.path.join(os.path.dirname(__file__), '../fake_data.xlsx')
    )

    # Pass --loop to run continuously, otherwise runs once
    if "--loop" in sys.argv:
        interval = int(os.getenv("SYNC_INTERVAL_SECONDS", "5"))
        run_loop(path, interval)
    else:
        run_once(path)
