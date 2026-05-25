"""
excel_push.py — runs on the Work PC.

Reads live data from a running Excel instance via xlwings and POSTs it
to the Mini PC's FastAPI /internal/push endpoint.

Modes:
    Default (production): connects to the running Excel process via xlwings.
    --file path/to/file.xlsx: reads from a static file via openpyxl (dev/testing).

Usage:
    python excel_push.py                        # xlwings, run once
    python excel_push.py --loop                 # xlwings, run continuously
    python excel_push.py --file data.xlsx       # openpyxl, run once
    python excel_push.py --file data.xlsx --loop  # openpyxl, run continuously

Config (via .env in this directory):
    MINIPC_API_URL      = https://your-service-name.onrender.com
    SYNC_API_KEY        = your-secret-key
    EXCEL_WORKBOOK_NAME = MyWorkbook.xlsx   (optional, matches partial name)
    PUSH_INTERVAL_SECONDS = 15
"""

import os
import sys
import json
import time
import logging
import argparse
import secrets
from datetime import datetime, timezone
from logging.handlers import RotatingFileHandler
from collections import deque
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeout

import requests
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))

# ── Config ────────────────────────────────────────────────────────────────────

MINIPC_API_URL       = os.getenv("MINIPC_API_URL", "http://localhost:8000")
SYNC_API_KEY         = os.getenv("SYNC_API_KEY", "")
EXCEL_WORKBOOK_NAME  = os.getenv("EXCEL_WORKBOOK_NAME", "")   # partial match, optional
PUSH_INTERVAL        = int(os.getenv("PUSH_INTERVAL_SECONDS", "15"))
EXCLUDED_SHEETS      = {s.strip() for s in os.getenv("EXCLUDED_SHEETS", "").split(",") if s.strip()}
MAX_RETRY_QUEUE      = 5   # max failed payloads to hold in memory

# ── Logging ───────────────────────────────────────────────────────────────────

log_path = os.path.join(os.path.dirname(__file__), "excel_push.log")
logger = logging.getLogger("excel_push")
logger.setLevel(logging.DEBUG)

file_handler = RotatingFileHandler(log_path, maxBytes=2_000_000, backupCount=3)
file_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)-8s %(message)s"))
stream_handler = logging.StreamHandler()
stream_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)-8s %(message)s"))
stream_handler.stream = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)
logger.addHandler(file_handler)
logger.addHandler(stream_handler)

# ── Retry queue ───────────────────────────────────────────────────────────────

retry_queue: deque = deque(maxlen=MAX_RETRY_QUEUE)


# ── Excel reading ─────────────────────────────────────────────────────────────

def serialize_value(val):
    """Convert cell values to JSON-serializable types."""
    if val is None:
        return None
    if hasattr(val, "isoformat"):   # datetime / date
        return val.isoformat()
    if isinstance(val, float) and val != val:  # NaN
        return None
    return val


def _is_title_row(row: list) -> bool:
    """A table-start row: the first non-None cell within the first 10 columns
    starts with ##. Cells beyond column 10 (e.g. RTD reference cells) are ignored.
    """
    for i, v in enumerate(row):
        if i > 10:
            break
        if v is None:
            continue
        return isinstance(v, str) and v.strip().startswith("##")
    return False


def _is_stop_row(row: list) -> bool:
    """A table-stop row: the first non-None cell within the first 10 columns
    starts with # but NOT ##. Cells beyond column 10 (e.g. RTD reference cells)
    are ignored so they don't prevent the stop from being detected.
    """
    for i, v in enumerate(row):
        if i > 10:
            break
        if v is None:
            continue
        if not isinstance(v, str):
            return False
        stripped = v.strip()
        return stripped.startswith("#") and not stripped.startswith("##")
    return False


def _header_columns(row: list) -> list[str]:
    headers = []
    for v in row:
        if v is None:
            break
        if isinstance(v, str) and v.strip():
            headers.append(v.strip())
    return headers


def _extract_title(row: list) -> str:
    val = next(v for v in row if v is not None)
    return str(val).strip().lstrip("#").strip()


# ── Cell formatting helpers ───────────────────────────────────────────────────

def _fmt_openpyxl(cell) -> dict | None:
    fmt = {}
    try:
        if cell.font and cell.font.bold:
            fmt['bold'] = True
    except Exception:
        pass
    try:
        fill = cell.fill
        if fill.patternType == 'solid':
            fc = fill.fgColor
            if fc.type == 'rgb' and len(fc.rgb) == 8 and fc.rgb[:2] != '00':
                fmt['bg'] = f"#{fc.rgb[2:]}"
    except Exception:
        pass
    try:
        fc = cell.font.color
        if fc and fc.type == 'rgb' and len(fc.rgb) == 8 and fc.rgb[:2] != '00':
            fmt['color'] = f"#{fc.rgb[2:]}"
    except Exception:
        pass
    return fmt if fmt else None


def _colorref_to_hex(colorref: int) -> str:
    r = colorref & 0xFF
    g = (colorref >> 8) & 0xFF
    b = (colorref >> 16) & 0xFF
    return f"#{r:02X}{g:02X}{b:02X}"


def _fmt_xlwings_cell(cell) -> dict | None:
    fmt = {}
    try:
        if cell.api.Font.Bold:
            fmt['bold'] = True
    except Exception:
        pass
    try:
        bg = cell.color  # (R, G, B) tuple or None = no fill
        if bg is not None:
            fmt['bg'] = f"#{bg[0]:02X}{bg[1]:02X}{bg[2]:02X}"
    except Exception:
        pass
    try:
        fc = cell.api.Font.Color  # COLORREF int
        if fc != 0:  # 0 = black (default)
            fmt['color'] = _colorref_to_hex(fc)
    except Exception:
        pass
    return fmt if fmt else None


def _read_fmt_xlwings(sheet) -> list[list] | None:
    try:
        used = sheet.used_range
        first_row = used.row
        first_col = used.column
        nrows = used.rows.count
        ncols = used.columns.count
        fmt_rows = []
        for r in range(first_row, first_row + nrows):
            row_fmts = []
            for c in range(first_col, first_col + ncols):
                row_fmts.append(_fmt_xlwings_cell(sheet.cells(r, c)))
            fmt_rows.append(row_fmts)
        return fmt_rows
    except Exception as e:
        logger.debug(f"Could not read cell formatting: {e}")
        return None


def parse_tables_from_rows(raw_rows: list[list], sheet_name: str, raw_fmts: list[list] | None = None) -> list[dict]:
    """
    Parse raw rows from one Excel sheet into one or more table dicts.

    Multi-table mode is activated when any ## marker row is found:
      - ## row  → flush current table, start a new named table.
      - #  row  → flush current table, discard this row and everything after
                  it until the next ## row.  Use it to close a table before
                  summary/decorative rows, or as a sheet-level title to skip.
      - blank   → always ignored.

    If no ## markers exist the whole sheet is treated as one table (backward
    compatible): first non-empty row = headers, rest = data.
    """
    has_markers = any(_is_title_row(r) for r in raw_rows)

    if not has_markers:
        non_empty_idx = [i for i, r in enumerate(raw_rows) if any(v is not None for v in r)]
        if not non_empty_idx:
            return []
        headers = _header_columns(raw_rows[non_empty_idx[0]])
        if not headers:
            return []
        data_rows = [raw_rows[i] for i in non_empty_idx[1:]]
        fmts = [raw_fmts[i] for i in non_empty_idx[1:]] if raw_fmts else None
        rows_serialized = _build_rows(headers, data_rows, fmts)
        logger.debug(f"  '{sheet_name}' (no ## markers): {len(rows_serialized)} rows, {len(headers)} cols")
        return [{"name": sheet_name, "source_sheet": sheet_name, "columns": headers, "rows": rows_serialized}]

    tables = []
    current_title: str | None = None
    current_headers: list | None = None
    current_data: list = []
    current_fmts: list = []

    for row_idx, row in enumerate(raw_rows):
        if _is_title_row(row):                              # ## → new table
            _flush_table(tables, current_title, current_headers, current_data,
                         current_fmts if raw_fmts else None)
            current_title = _extract_title(row) or f"Tabla {len(tables) + 1}"
            current_headers = None
            current_data = []
            current_fmts = []
        elif _is_stop_row(row):                             # #  → close table
            _flush_table(tables, current_title, current_headers, current_data,
                         current_fmts if raw_fmts else None)
            current_title = None
            current_headers = None
            current_data = []
            current_fmts = []
        elif all(v is None for v in row):
            continue                                        # blank → skip
        elif current_title is not None and current_headers is None:
            headers = _header_columns(row)
            if headers:
                current_headers = headers
        elif current_headers is not None:
            current_data.append(row)
            if raw_fmts:
                current_fmts.append(raw_fmts[row_idx] if row_idx < len(raw_fmts) else None)

    _flush_table(tables, current_title, current_headers, current_data,
                 current_fmts if raw_fmts else None)
    return tables


def _build_rows(headers: list, data_rows: list, fmt_rows: list | None = None) -> list[dict]:
    result = []
    for i, row in enumerate(data_rows):
        d = {headers[j]: serialize_value(row[j] if j < len(row) else None)
             for j in range(len(headers))}
        if fmt_rows and i < len(fmt_rows) and fmt_rows[i]:
            row_fmt = fmt_rows[i]
            fmt = {headers[j]: row_fmt[j]
                   for j in range(len(headers))
                   if j < len(row_fmt) and row_fmt[j]}
            if fmt:
                d['__fmt__'] = fmt
        result.append(d)
    return result


def _flush_table(tables: list, title: str | None, headers: list | None, data: list,
                 fmts: list | None = None):
    if not title or not headers:
        if title and not headers:
            logger.warning(f"  Table '{title}' has no header row — skipped.")
        return
    rows_serialized = _build_rows(headers, data, fmts)
    logger.debug(f"  Table '{title}': {len(rows_serialized)} rows, {len(headers)} cols")
    tables.append({"name": title, "columns": headers, "rows": rows_serialized})


def read_via_xlwings(only_sheet: list[str] | None = None) -> list[dict] | None:
    """
    Connect to the running Excel process via xlwings and read all sheets.
    Returns a flat list of table dicts (one per ## section, or one per sheet).
    """
    try:
        import xlwings as xw
    except ImportError:
        logger.error("xlwings is not installed. Run: pip install xlwings")
        return None

    try:
        app = xw.apps.active
        if app is None:
            raise RuntimeError("No active Excel app found.")
    except Exception:
        logger.warning("Excel is not running — skipping this cycle.")
        return None

    try:
        if EXCEL_WORKBOOK_NAME:
            wb = next(
                (b for b in app.books if EXCEL_WORKBOOK_NAME.lower() in b.name.lower()),
                None
            )
            if wb is None:
                logger.warning(
                    f"Workbook matching '{EXCEL_WORKBOOK_NAME}' not found. "
                    f"Open workbooks: {[b.name for b in app.books]}"
                )
                return None
        else:
            if not app.books:
                logger.warning("No workbooks open in Excel.")
                return None
            wb = app.books[0]
    except Exception as e:
        logger.error(f"Error accessing workbooks: {e}")
        return None

    all_tables = []
    try:
        for sheet in wb.sheets:
            sheet_name = sheet.name

            if sheet_name in EXCLUDED_SHEETS:
                logger.debug(f"Skipping '{sheet_name}': excluded.")
                continue

            if only_sheet and sheet_name not in only_sheet:
                logger.debug(f"Skipping '{sheet_name}': not selected.")
                continue

            raw = sheet.used_range.value
            if raw is None:
                logger.debug(f"Skipping '{sheet_name}': empty.")
                continue
            if not isinstance(raw, list):
                raw = [[raw]]
            elif raw and not isinstance(raw[0], list):
                raw = [raw]

            fmt_rows = _read_fmt_xlwings(sheet)
            logger.debug(f"Sheet '{sheet_name}':")
            tables = parse_tables_from_rows(raw, sheet_name, fmt_rows)
            for t in tables:
                t["source_sheet"] = sheet_name
            all_tables.extend(tables)

    except Exception as e:
        logger.error(f"Error reading workbook '{wb.name}': {e}")
        return None

    return all_tables


def read_via_openpyxl(file_path: str, only_sheet: list[str] | None = None) -> list[dict] | None:
    """Read from a static .xlsx file. Used in dev/testing mode."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except FileNotFoundError:
        logger.error(f"File not found: {file_path}")
        return None
    except Exception as e:
        logger.error(f"Could not open file '{file_path}': {e}")
        return None

    all_tables = []
    for sheet_name in wb.sheetnames:
        if sheet_name in EXCLUDED_SHEETS:
            logger.debug(f"Skipping '{sheet_name}': excluded.")
            continue

        if only_sheet and sheet_name not in only_sheet:
            logger.debug(f"Skipping '{sheet_name}': not selected.")
            continue

        ws = wb[sheet_name]
        ws_rows = list(ws.iter_rows())
        raw_rows = [[cell.value for cell in row] for row in ws_rows]
        fmt_rows = [[_fmt_openpyxl(cell) for cell in row] for row in ws_rows]

        if not raw_rows:
            logger.debug(f"Skipping '{sheet_name}': empty.")
            continue

        logger.debug(f"Sheet '{sheet_name}':")
        tables = parse_tables_from_rows(raw_rows, sheet_name, fmt_rows)
        for t in tables:
            t["source_sheet"] = sheet_name
        all_tables.extend(tables)

    return all_tables


# ── HTTP push ─────────────────────────────────────────────────────────────────

def push_to_api(sheets: list[dict]) -> bool:
    """
    POST sheet data to the Mini PC's FastAPI endpoint.
    Returns True on success.
    """
    if not SYNC_API_KEY:
        logger.error("SYNC_API_KEY is not set in .env — cannot push.")
        return False

    url = f"{MINIPC_API_URL.rstrip('/')}/internal/push"
    payload = {"sheets": sheets, "pushed_at": datetime.now(timezone.utc).isoformat()}
    headers = {"X-API-Key": SYNC_API_KEY, "Content-Type": "application/json"}

    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=10)
        if resp.status_code == 200:
            result = resp.json()
            logger.info(
                f"Push OK — {result.get('sheets_synced', '?')} sheets, "
                f"{result.get('total_rows', '?')} rows"
            )
            return True
        elif resp.status_code == 401:
            logger.error("Push rejected: invalid API key.")
            return False
        else:
            logger.warning(f"Push failed: HTTP {resp.status_code} — {resp.text[:200]}")
            return False
    except requests.exceptions.ConnectionError:
        logger.warning(f"Mini PC unreachable at {MINIPC_API_URL} — queuing for retry.")
        return False
    except requests.exceptions.Timeout:
        logger.warning("Push timed out — queuing for retry.")
        return False
    except Exception as e:
        logger.error(f"Unexpected push error: {e}")
        return False


def flush_retry_queue():
    """Try to send any previously failed payloads."""
    if not retry_queue:
        return
    logger.info(f"Retrying {len(retry_queue)} queued payload(s)...")
    still_failing = []
    while retry_queue:
        queued = retry_queue.popleft()
        if not push_to_api(queued):
            still_failing.append(queued)
    for item in still_failing:
        retry_queue.append(item)


# ── Main cycle ────────────────────────────────────────────────────────────────

def read_with_timeout(file_path: str | None, timeout: int = 15, only_sheet: list[str] | None = None) -> list[dict] | None:
    """Run the Excel read in a thread with a timeout. Returns None if Excel is busy."""
    fn = (lambda: read_via_openpyxl(file_path, only_sheet)) if file_path else (lambda: read_via_xlwings(only_sheet))
    with ThreadPoolExecutor(max_workers=1) as executor:
        future = executor.submit(fn)
        try:
            return future.result(timeout=timeout)
        except FuturesTimeout:
            logger.warning("Excel read timed out (Excel may be busy or showing a dialog) — skipping cycle.")
            return None


def run_cycle(file_path: str | None, only_sheet: list[str] | None = None):
    """Read Excel and push. Queues on failure."""
    logger.info("--- Sync cycle starting ---")

    t0 = time.perf_counter()
    sheets = read_with_timeout(file_path, only_sheet=only_sheet)
    read_ms = (time.perf_counter() - t0) * 1000
    logger.info(f"Excel read: {read_ms:.0f}ms")

    if sheets is None:
        logger.warning("No data read — skipping push.")
        return

    if not sheets:
        logger.warning("All sheets were empty — skipping push.")
        return

    flush_retry_queue()

    t1 = time.perf_counter()
    success = push_to_api(sheets)
    push_ms = (time.perf_counter() - t1) * 1000
    logger.info(f"HTTP push: {push_ms:.0f}ms | total cycle: {(read_ms + push_ms):.0f}ms")

    if not success:
        retry_queue.append(sheets)
        logger.warning(f"Queued for retry. Queue size: {len(retry_queue)}")


def run_loop(file_path: str | None, interval: int, only_sheet: list[str] | None = None):
    mode = f"file: {file_path}" if file_path else "xlwings (live Excel)"
    logger.info(f"Starting push loop — interval: {interval}s — mode: {mode}")
    while True:
        run_cycle(file_path, only_sheet=only_sheet)
        logger.info(f"Sleeping {interval}s...\n")
        time.sleep(interval)


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Excel → Mini PC push script")
    parser.add_argument("--file", metavar="PATH", help="Use openpyxl on a file instead of win32com")
    parser.add_argument("--loop", action="store_true", help="Run continuously")
    parser.add_argument("--sheet", metavar="NAME", nargs="+", help="Only push these sheets (space-separated exact names)")
    args = parser.parse_args()

    if args.loop:
        run_loop(args.file, PUSH_INTERVAL, only_sheet=args.sheet)
    else:
        run_cycle(args.file, only_sheet=args.sheet)
